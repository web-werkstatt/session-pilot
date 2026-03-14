"""
Export-Formate fuer Claude Sessions: JSON, XLSX, MD, HTML, TXT
"""
import json
import io
from datetime import datetime
from services.db_service import execute


def get_session_with_messages(session_uuid):
    """Laedt Session mit allen Messages"""
    session = execute(
        "SELECT * FROM sessions WHERE session_uuid = %s", (session_uuid,), fetchone=True
    )
    if not session:
        return None

    messages = execute(
        "SELECT * FROM messages WHERE session_id = %s ORDER BY timestamp ASC",
        (session["id"],), fetch=True
    )
    return dict(session), [dict(m) for m in messages]


def format_duration(ms):
    """Formatiert Millisekunden in lesbare Dauer"""
    if not ms:
        return "0s"
    seconds = ms // 1000
    if seconds < 60:
        return f"{seconds}s"
    minutes = seconds // 60
    secs = seconds % 60
    if minutes < 60:
        return f"{minutes}m {secs}s"
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours}h {mins}m"


def format_tokens(n):
    """Formatiert Token-Zahlen"""
    if not n:
        return "0"
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if n >= 1000:
        return f"{n / 1000:.1f}K"
    return str(n)


def _serialize_session(session):
    """Konvertiert Session-Dict fuer JSON-Serialisierung"""
    result = {}
    for k, v in session.items():
        if isinstance(v, datetime):
            result[k] = v.isoformat()
        else:
            result[k] = v
    return result


def _serialize_message(msg):
    """Konvertiert Message-Dict fuer JSON-Serialisierung"""
    result = {}
    for k, v in msg.items():
        if isinstance(v, datetime):
            result[k] = v.isoformat()
        else:
            result[k] = v
    return result


def export_json(session_uuid):
    """Exportiert Session als JSON"""
    data = get_session_with_messages(session_uuid)
    if not data:
        return None, None
    session, messages = data

    export = {
        "session": _serialize_session(session),
        "messages": [_serialize_message(m) for m in messages]
    }
    content = json.dumps(export, indent=2, ensure_ascii=False, default=str)
    filename = f"session_{session.get('slug') or session_uuid[:8]}.json"
    return content.encode("utf-8"), filename


def export_markdown(session_uuid):
    """Exportiert Session als Markdown"""
    data = get_session_with_messages(session_uuid)
    if not data:
        return None, None
    session, messages = data

    lines = []
    lines.append(f"# Claude Session: {session.get('project_name', 'Unbekannt')}")
    lines.append("")
    lines.append(f"- **Account:** {session.get('account', '-')}")
    lines.append(f"- **Datum:** {session.get('started_at', '-')}")
    lines.append(f"- **Dauer:** {format_duration(session.get('duration_ms'))}")
    lines.append(f"- **Model:** {session.get('model', '-')}")
    lines.append(f"- **Tokens:** {format_tokens(session.get('total_input_tokens'))} in / {format_tokens(session.get('total_output_tokens'))} out")
    lines.append(f"- **Branch:** {session.get('git_branch', '-')}")
    lines.append(f"- **CWD:** {session.get('cwd', '-')}")
    lines.append("")
    lines.append("---")
    lines.append("")

    for msg in messages:
        if msg["type"] == "system":
            lines.append(f"*--- {msg.get('content', '')} ---*")
            lines.append("")
            continue

        role_label = "User" if msg["type"] == "user" else "Assistant"
        lines.append(f"## {role_label}")
        lines.append("")

        content = msg.get("content", "")
        if content:
            lines.append(content)
        lines.append("")

    content = "\n".join(lines)
    filename = f"session_{session.get('slug') or session_uuid[:8]}.md"
    return content.encode("utf-8"), filename


def export_html(session_uuid):
    """Exportiert Session als standalone Dark-Theme HTML"""
    data = get_session_with_messages(session_uuid)
    if not data:
        return None, None
    session, messages = data

    try:
        import markdown as md_lib
        def md(text):
            return md_lib.markdown(text, extensions=["fenced_code", "tables"])
    except ImportError:
        def md(text):
            import html
            return f"<pre>{html.escape(text)}</pre>"

    msg_html = []
    for msg in messages:
        if msg["type"] == "system":
            msg_html.append(f'<div class="system">{msg.get("content", "")}</div>')
            continue

        cls = "user" if msg["type"] == "user" else "assistant"
        content = msg.get("content", "")
        rendered = md(content) if content else ""
        msg_html.append(f'<div class="msg {cls}"><div class="role">{cls.title()}</div>{rendered}</div>')

    html_content = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<title>Session: {session.get('project_name', '')}</title>
<style>
body {{ font-family: 'Segoe UI', sans-serif; background: #1e1e1e; color: #ddd; max-width: 900px; margin: 0 auto; padding: 20px; }}
h1 {{ color: #4fc3f7; }}
.meta {{ background: #2d2d2d; padding: 15px; border-radius: 8px; margin-bottom: 20px; }}
.meta span {{ margin-right: 20px; color: #aaa; }}
.msg {{ margin: 15px 0; padding: 15px; border-radius: 8px; }}
.msg.user {{ background: #1a3a5c; border-left: 3px solid #4fc3f7; }}
.msg.assistant {{ background: #2d2d2d; border-left: 3px solid #66bb6a; }}
.role {{ font-weight: bold; margin-bottom: 8px; color: #888; text-transform: uppercase; font-size: 12px; }}
.system {{ text-align: center; color: #666; font-size: 12px; margin: 10px 0; }}
pre {{ background: #1a1a1a; padding: 12px; border-radius: 5px; overflow-x: auto; }}
code {{ font-family: 'Fira Code', monospace; font-size: 13px; }}
@media print {{ body {{ background: white; color: black; }} .msg {{ border: 1px solid #ddd; }} }}
</style>
</head>
<body>
<h1>{session.get('project_name', 'Session')}</h1>
<div class="meta">
<span>Account: {session.get('account', '-')}</span>
<span>Dauer: {format_duration(session.get('duration_ms'))}</span>
<span>Model: {session.get('model', '-')}</span>
<span>Tokens: {format_tokens(session.get('total_input_tokens'))}in / {format_tokens(session.get('total_output_tokens'))}out</span>
</div>
{''.join(msg_html)}
</body>
</html>"""

    filename = f"session_{session.get('slug') or session_uuid[:8]}.html"
    return html_content.encode("utf-8"), filename


def export_xlsx(session_uuid):
    """Exportiert Session als XLSX"""
    data = get_session_with_messages(session_uuid)
    if not data:
        return None, None
    session, messages = data

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # Sheet 1: Metadaten
    ws1 = wb.active
    ws1.title = "Session"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

    meta_fields = [
        ("Feld", "Wert"),
        ("Session-UUID", session.get("session_uuid", "")),
        ("Account", session.get("account", "")),
        ("Projekt", session.get("project_name", "")),
        ("CWD", session.get("cwd", "")),
        ("Branch", session.get("git_branch", "")),
        ("Model", session.get("model", "")),
        ("Version", session.get("claude_version", "")),
        ("Slug", session.get("slug", "")),
        ("Start", str(session.get("started_at", ""))),
        ("Ende", str(session.get("ended_at", ""))),
        ("Dauer", format_duration(session.get("duration_ms"))),
        ("User-Nachrichten", session.get("user_message_count", 0)),
        ("Assistant-Nachrichten", session.get("assistant_message_count", 0)),
        ("Input-Tokens", session.get("total_input_tokens", 0)),
        ("Output-Tokens", session.get("total_output_tokens", 0)),
    ]
    for row_idx, (field, value) in enumerate(meta_fields, 1):
        ws1.cell(row=row_idx, column=1, value=field)
        ws1.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws1.cell(row=row_idx, column=1).font = header_font
            ws1.cell(row=row_idx, column=1).fill = header_fill
            ws1.cell(row=row_idx, column=2).font = header_font
            ws1.cell(row=row_idx, column=2).fill = header_fill

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 60

    # Sheet 2: Messages
    ws2 = wb.create_sheet("Messages")
    headers = ["#", "Typ", "Zeitstempel", "Model", "Input-Tokens", "Output-Tokens", "Inhalt"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for idx, msg in enumerate(messages, 1):
        content = msg.get("content", "")
        # Excel-Zellenlimit: max 32767 Zeichen
        if len(content) > 32000:
            content = content[:32000] + "... (gekuerzt)"
        ws2.cell(row=idx + 1, column=1, value=idx)
        ws2.cell(row=idx + 1, column=2, value=msg.get("type", ""))
        ws2.cell(row=idx + 1, column=3, value=str(msg.get("timestamp", "")))
        ws2.cell(row=idx + 1, column=4, value=msg.get("model", ""))
        ws2.cell(row=idx + 1, column=5, value=msg.get("input_tokens", 0))
        ws2.cell(row=idx + 1, column=6, value=msg.get("output_tokens", 0))
        ws2.cell(row=idx + 1, column=7, value=content)

    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["G"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"session_{session.get('slug') or session_uuid[:8]}.xlsx"
    return buf.getvalue(), filename


def export_txt(session_uuid):
    """Exportiert Session als einfacher Text"""
    data = get_session_with_messages(session_uuid)
    if not data:
        return None, None
    session, messages = data

    lines = []
    lines.append(f"Claude Session: {session.get('project_name', 'Unbekannt')}")
    lines.append(f"Account: {session.get('account', '-')} | Datum: {session.get('started_at', '-')}")
    lines.append(f"Dauer: {format_duration(session.get('duration_ms'))} | Model: {session.get('model', '-')}")
    lines.append("=" * 80)
    lines.append("")

    for msg in messages:
        if msg["type"] == "system":
            lines.append(f"--- {msg.get('content', '')} ---")
            lines.append("")
            continue

        role = "USER" if msg["type"] == "user" else "ASSISTANT"
        lines.append(f"[{role}]")
        content = msg.get("content", "")
        if content:
            lines.append(content)
        lines.append("")
        lines.append("-" * 40)
        lines.append("")

    content = "\n".join(lines)
    filename = f"session_{session.get('slug') or session_uuid[:8]}.txt"
    return content.encode("utf-8"), filename
